import openpyxl

path = '/Users/yangqihe/Desktop/2017a.xlsx'
wb = openpyxl.load_workbook(path)

sheet2 = wb['Sheet2']
sheet3 = wb['Sheet3']

ws = wb.active
columns2 = sheet2['A:KA']
columnCount = len(columns2)
rowColumn = columns2[0]
rowsCount = len(rowColumn)
print(rowsCount)

columns3 = sheet3['A:KA']

allArray = []

for rIndex in range(rowsCount):
    if rIndex < 3:
        if rIndex > 0:
            rowCellValue = rowColumn[rIndex].value
            currentArray = []
            for cIndex in range(columnCount):
                if cIndex > 0:
                    column = columns2[cIndex]
                    columnCellValue = column[rIndex].value
                    arrayValue = ''
                    if columnCellValue is None or str(columnCellValue).strip() == '':
                        arrayValue = rowCellValue + "," + column[0].value + "," + '0'
                    else:
                        arrayValue = rowCellValue + "," + column[0].value + "," + str(columnCellValue).strip()
                    currentArray.append(arrayValue)
            allArray.append(currentArray)
    else:
        break

indexI = 0
for allIndex in range(len(allArray)):
    currentArray = allArray[allIndex]
    for i in range(len(currentArray)):
        currentArray_ = currentArray.copy()
        temp = currentArray_[i]
        currentArray_.remove(temp)
        temps = temp.split(',')
        currentArray_.insert(0, temps[2])
        currentArray_.insert(0, temps[0]+temps[1])
        indexI += 1
        indexJ = 0
        for j in range(len(currentArray_)):
            indexJ += 1
            print(indexI, ",", indexJ)
            cellValue = currentArray_[j]
            if indexJ > 2:
                cellValues = cellValue.split(',')
                sheet3.cell(row=indexI, column=indexJ).value = cellValues[2]
            else:
                sheet3.cell(row=indexI, column=indexJ).value = cellValue

wb.save(path)
