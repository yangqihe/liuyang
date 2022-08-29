import openpyxl

path = '/Users/yangqihe/Desktop/π.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active
sheet1 = wb['Sheet1']
columns = sheet1['A:D']
columnA = columns[0]
columnB = columns[1]
columnC = columns[2]
columnD = columns[3]
print("总数: " + str(len(columnA)))


for i in range(len(columnA)):
    # if i > 200:
    #     break
    print('........'+str(i)+'.........')
    rowDestinationCityI = columnA[i].value
    rowOriginCityI = columnB[i].value
    for j in range(len(columnA)):
        rowDestinationCityJ = columnA[j].value
        rowOriginCityJ = columnB[j].value

        # DO == OD
        if rowDestinationCityI + rowOriginCityI == rowOriginCityJ + rowDestinationCityJ:
            sheet1.cell(row=i+1, column=5).value = columnA[j].value
            sheet1.cell(row=i+1, column=6).value = columnB[j].value
            sheet1.cell(row=i+1, column=7).value = columnC[j].value
            sheet1.cell(row=i+1, column=8).value = columnD[j].value

        # D == O
        if rowDestinationCityI == rowDestinationCityJ == rowOriginCityJ:
            sheet1.cell(row=i + 1, column=9).value = columnA[j].value
            sheet1.cell(row=i + 1, column=10).value = columnB[j].value
            sheet1.cell(row=i + 1, column=11).value = columnC[j].value
            sheet1.cell(row=i + 1, column=12).value = columnD[j].value

        # O == D
        if rowOriginCityI == rowDestinationCityJ == rowOriginCityJ:
            sheet1.cell(row=i + 1, column=13).value = columnA[j].value
            sheet1.cell(row=i + 1, column=14).value = columnB[j].value
            sheet1.cell(row=i + 1, column=15).value = columnC[j].value
            sheet1.cell(row=i + 1, column=16).value = columnD[j].value

wb.save(path)
